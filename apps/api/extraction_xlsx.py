"""XLSX extraction (roadmap step 092) -- registered into
extraction.py:HANDLERS under "xlsx".

openpyxl (MIT), read_only=True (streams rows instead of loading the
whole workbook into memory -- consistent with this project's general
memory-safety stance, e.g. validation.py's chunked upload reads) and
data_only=True (reads each cell's last-calculated value rather than its
formula string -- a client uploading a spreadsheet wants the numbers
it shows, not "=SUM(A1:A9)").

Each sheet becomes a "# {sheet name}" heading followed by its rows as
one markdown table, treating row 1 as the header -- the common shape for
real spreadsheet exports, not a guarantee openpyxl or the xlsx format
itself makes. Cell values are booleans/numbers/dates/None, not strings
(unlike every other extractor's table cells so far) -- rows_to_markdown
already stringifies non-None cells and treats only None as an empty
cell, not falsy values like 0 or False, so no per-cell conversion is
needed here beyond what that shared helper already does.
"""

import io

import openpyxl

from extraction_tables import rows_to_markdown


def extract_xlsx(content: bytes) -> str:
    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        sheet_texts: list[str] = []
        for name in workbook.sheetnames:
            sheet = workbook[name]
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            markdown_table = rows_to_markdown(rows)
            if markdown_table:
                sheet_texts.append(f"# {name}\n\n{markdown_table}")
        return "\n\n".join(sheet_texts)
    finally:
        workbook.close()


def extract_xlsx_metadata(content: bytes) -> dict[str, object]:
    """Raw workbook.properties values, uncleaned -- see
    extraction_metadata.py for why cleaning is centralized there instead
    of here. Note the field is `creator`, not `author` like docx/pptx --
    normalized to a common "author" key here so callers don't need to
    know the per-format field name."""
    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    try:
        props = workbook.properties
        return {
            "title": props.title,
            "author": props.creator,
            "created_at": props.created.isoformat() if props.created else None,
            "modified_at": props.modified.isoformat() if props.modified else None,
            "language": props.language,
        }
    finally:
        workbook.close()
